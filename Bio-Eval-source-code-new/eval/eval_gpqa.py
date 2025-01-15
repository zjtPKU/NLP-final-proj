import json
import re
from collections import Counter
import argparse
import os
from prettytable import PrettyTable
import pandas as pd
from typing import Dict, List
from openpyxl.styles import PatternFill, Font, Alignment
from tqdm import tqdm
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont


def normalize_response(response: str) -> str:
    """
    Normalize the response by removing markdown and LaTeX formatting that may prevent a match.
    """

    return (
        response.replace("**", "")
        .replace("$\\boxed{", "")
        .replace("}$", "")
        .replace("\\$", "")
        .replace("$\\text{", "")
        .replace("$", "")
        .replace("\\mathrm{", "")
        .replace("\\{", "")
        .replace("\\text", "")
        .replace("\\(", "")
        .replace("\\mathbf{", "")
        .replace("{", "")
        .replace("\\boxed", "")
    )

def extract_option_labels(text, options='ABCDEFGHIJ'):
    if isinstance(text, dict):
        return 'error'
    
    # 标准化文本
    text = normalize_response(text)
    
    # 去掉尾部换行符并获取最后一行
    text = text.rstrip()
    last_line = text.split('\n')[-1]
    
    option_str = ''.join([chr(65 + i) for i in range(len(options))]) if options else 'ABCDEFGHIJ'
    
    patterns = [
        # 优先匹配不带括号的格式
        # 匹配 "The answer to the question is: A"
        f'[Tt]he answer to the question is:?\s*([{option_str}])\s*.*',
        
        # 匹配 "The correct answer option is: A"
        f'[Tt]he correct answer option is:?\s*([{option_str}])\s*.*',
        
        # 匹配 "The correct option is: A" 
        f'[Tt]he correct option is:?\s*([{option_str}])\s*.*',
        
        # 匹配 "The correct answer is: A"
        f'[Tt]he correct answer is:?\s*([{option_str}])\s*.*',
        
        # 匹配 "The answer is: A"
        f'[Tt]he answer is:?\s*([{option_str}])\s*.*',
        
        # 匹配 "ANSWER: A"
        f'(?i)ANSWER\s*:\s*([{option_str}])\s*.*',
        
        # 匹配 "A: " 或 "A."
        f'([{option_str}])\s*[：:.]',
        
        # 然后再匹配带括号的格式
        f'[Tt]he answer to the question is:?\s*\(?([{option_str}])\)?\s*.*',
        
        # 匹配 "The correct answer option is: boxed(A)"
        f'[Tt]he correct answer option is:?.*?boxed{{\(?([{option_str}])\)?}}\s*.*',
        
        # 匹配 "The correct option is: boxed(A)"
        f'[Tt]he correct option is:?.*?boxed{{\(?([{option_str}])\)?}}\s*.*',
        
        # 匹配 "The correct answer is: boxed(A)"
        f'[Tt]he correct answer is:?.*?boxed{{\(?([{option_str}])\)?}}\s*.*',

        # 匹配 "The correct answer is option: (A)"
        f'[Tt]he correct answer is option:?\s*\(?([{option_str}])\)?\s*.*',

        # 匹配 "The correct answer is: (A)"
        f'[Tt]he correct answer is:?\s*\(?([{option_str}])\)?\s*.*',
        
        # 匹配 "The answer is option: (A)"
        f'[Tt]he answer is option:?\s*\(?([{option_str}])\)?\s*.*',

        # 匹配 "The answer is: (A)"
        f'[Tt]he answer is:?\s*\(?([{option_str}])\)?\s*.*',

        # 匹配 "ANSWER: (A)"
        f'(?i)ANSWER\s*:\s*\(?([{option_str}])\)?\s*.*',
        
        # 匹配 "(A) " 或 "(A)."
        f'\(?([{option_str}])\)?\s*(?:[。，,：:\.$])?',
        
        # 匹配 "(A)" 或 "(A)"
        f'\(?([{option_str}])\)?\s*.*',

        # 匹配 "A: " 或 "A."
        f'([{option_str}])\s*:',
    ]
    
    # 首先尝试在最后一行中匹配
    for pattern in patterns:
        match = re.search(pattern, last_line)
        if match:
            return match.group(1)
    
    # 如果最后一行没有匹配到,则尝试在整个文本中匹配
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    
    return None

def calculate_accuracy(file_path, save_dir):
    data = []
    acc = 0
    count = 0
    err = 0
    miss = 0
    
    stats = {
        'overview': {},
        'category': {},
        'subcategory': {}
    }
    
    # 读取数据
    with open(file_path, "r") as file:
        for line in tqdm(file, desc=f"读取{os.path.basename(file_path)}中的数据", leave=False):
            data_ = json.loads(line)
            data.append(data_)
    
    if not data:
        print(f"Warning: No data found in {file_path}")
        return 0.0, 0.0, 0.0, stats
    
    # 处理每个样本
    for sample in tqdm(data, desc=f"处理{os.path.basename(file_path)}中的样本", leave=False):
        predict = extract_option_labels(sample["response"], 'ABCDEFGHIJ')
        sample["extracted_answer"] = predict
        
        # 获取所有层级的类别信息
        overview = sample.get("overview_category", "unknown")
        category = sample.get("category", "unknown")
        subcategory = sample.get("subcategory", "unknown")
        
        # 初始化各层级的统计
        for level, key in [
            ('overview', overview),
            ('category', f"{overview}/{category}"),
            ('subcategory', f"{overview}/{category}/{subcategory}")
        ]:
            if key not in stats[level]:
                stats[level][key] = {
                    "correct": 0, 
                    "total": 0, 
                    "miss": 0, 
                    "error": 0,
                    "overview_category": overview,
                    "category": category,
                    "subcategory": subcategory
                }
            
            stats[level][key]["total"] += 1
            # answer_letter = sample["answer_letter"]
            answer_letter = chr(sample["options"].index(sample["answer"]) + 65)
            # match = re.search(r'Answer: ([A-J])', sample["response_answer"])
            # if match:
            #     answer_letter = match.group(1)
            # else:
            #     continue
            
            if predict and answer_letter == predict:
                acc += 1
                sample["status"] = "correct"
                stats[level][key]["correct"] += 1
            elif predict == None or predict == "":
                miss += 1
                sample["status"] = "miss"
                stats[level][key]["miss"] += 1
            elif predict == 'error':
                err += 1
                sample["status"] = "error"
                stats[level][key]["error"] += 1
            else:
                sample["status"] = "incorrect"
            count += 1
    
    if count == 0:
        print(f"Warning: No valid samples found in {file_path}")
        return 0.0, 0.0, 0.0, stats
    
    accuracy = acc / count
    errors = err / count
    miss_rate = miss / count
    
    os.makedirs(save_dir, exist_ok=True)
    save_path = os.path.join(save_dir, os.path.basename(file_path))
    with open(save_path, "w") as file:
        for sample in data:
            json.dump(sample, file)
            file.write("\n")
    
    return accuracy, errors, miss_rate, stats

def create_excel_report_from_stats(model_results, hierarchy_stats, save_path):
    print("开始生成Excel报告...")
    rows = []
    
    # 按层级组织数据
    for overview in tqdm(sorted(hierarchy_stats['overview'].keys()), desc="处理overview层级"):
        overview_stats = hierarchy_stats['overview'][overview]
        overview_total = overview_stats['total']
        
        # 获取该overview下的所有category
        categories = [k for k in hierarchy_stats['category'].keys() 
                     if k.startswith(f"{overview}/")]
        
        for category_key in sorted(categories):
            category_stats = hierarchy_stats['category'][category_key]
            category = category_stats['category']
            category_total = category_stats['total']
            
            # 获取该category下的所有subcategory
            subcategories = [k for k in hierarchy_stats['subcategory'].keys() 
                           if k.startswith(f"{overview}/{category}/")]
            
            # 添加subcategory行
            for subcategory_key in sorted(subcategories):
                subcategory_stats = hierarchy_stats['subcategory'][subcategory_key]
                row_data = {
                    'Overview Category': overview,
                    'Category': category,
                    'Subcategory': f"{subcategory_stats['subcategory']} ({subcategory_stats['total']})"
                }
                
                # 添加每个模型的得分
                for model_name in model_results.keys():
                    for mode in model_results[model_name].keys():
                        stats = subcategory_stats['model_stats'].get(model_name, {}).get(mode, 
                                {'total': 0, 'correct': 0, 'error': 0, 'miss': 0})
                        row_data[f'{model_name}_{mode}'] = format_cell_value(stats)
                
                rows.append(row_data)
            
            # 添加category汇总行
            category_row = {
                'Overview Category': overview,
                'Category': f"{category} (Total: {category_total})",
                'Subcategory': ''
            }
            for model_name in model_results.keys():
                for mode in model_results[model_name].keys():
                    stats = category_stats['model_stats'].get(model_name, {}).get(mode, 
                            {'total': 0, 'correct': 0, 'error': 0, 'miss': 0})
                    category_row[f'{model_name}_{mode}'] = format_cell_value(stats)
            rows.append(category_row)
        
        # 添加overview汇总行
        overview_row = {
            'Overview Category': f"{overview} (Total: {overview_total})",
            'Category': '',
            'Subcategory': ''
        }
        for model_name in model_results.keys():
            for mode in model_results[model_name].keys():
                stats = overview_stats['model_stats'].get(model_name, {}).get(mode, 
                        {'total': 0, 'correct': 0, 'error': 0, 'miss': 0})
                overview_row[f'{model_name}_{mode}'] = format_cell_value(stats)
        rows.append(overview_row)

    # 创建DataFrame
    df = pd.DataFrame(rows)
    
    # 添加总体汇总行
    total_stats = {}
    for model_name in model_results.keys():
        for mode in model_results[model_name].keys():
            # 从每个模型的具体统计中获取数据
            model_stats = hierarchy_stats['overview']
            model_total = {
                'total': 0,
                'correct': 0,
                'error': 0,
                'miss': 0
            }
            
            # 累加该模型下所有overview类别的统计数据
            for category_stats in model_stats.values():
                if 'model_stats' in category_stats and model_name in category_stats['model_stats']:
                    stats = category_stats['model_stats'][model_name].get(mode, {})
                    model_total['total'] += stats.get('total', 0)
                    model_total['correct'] += stats.get('correct', 0)
                    model_total['error'] += stats.get('error', 0)
                    model_total['miss'] += stats.get('miss', 0)
            
            total_stats[f"{model_name}_{mode}"] = model_total
    
    total_row = {
        'Overview Category': f"Overall Total",
        'Category': '',
        'Subcategory': ''
    }
    
    # 为每个模型和模式添加对应的统计数据
    for col in df.columns[3:]:  # 跳过前三列
        if col in total_stats:
            total_row[col] = format_cell_value(total_stats[col])
        else:
            total_row[col] = '0%/0%/0%'
            
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    
    # 定义颜色和字体
    COLORS = {
        'acc': '000000',     # 黑色
        'error': 'FF6B6B',   # 柔和的红色
        'miss': '808080',    # 灰色
        'overview_total': 'FFFF00',  # 明黄色
        'category_total': 'FFFFD4',  # 淡黄色
        'final_total': 'FFB6C1'      # 粉红色
    }
    
    # 设置默认字体
    default_font = Font(name='微软雅黑')
    header_font = Font(name='微软雅黑', bold=True)
    
    # 设置富文本字体 - InlineFont只支持sz, b, i, u, strike, color属性
    default_inline_font = InlineFont()
    acc_inline_font = InlineFont(color=COLORS['acc'])
    error_inline_font = InlineFont(color=COLORS['error'])
    miss_inline_font = InlineFont(color=COLORS['miss'])
    
    # 保存到Excel
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Results')
        worksheet = writer.sheets['Results']
        
        # 设置所有单元格的默认字体为微软雅黑
        for row in worksheet.rows:
            for cell in row:
                cell.font = default_font
        
        # 设置单元格格式
        for row in worksheet.iter_rows(min_row=2):
            # 处理数据单元格
            for cell in row[3:]:  # 跳过前三列
                if cell.value and isinstance(cell.value, str) and '/' in cell.value:
                    try:
                        parts = cell.value.split('/')
                        if len(parts) == 3:
                            rich_text = CellRichText()
                            # 正确率（黑色）
                            rich_text.append(TextBlock(text=parts[0], font=acc_inline_font))
                            # 分隔符
                            rich_text.append(TextBlock(text="/", font=default_inline_font))
                            # 错误率（红色）
                            rich_text.append(TextBlock(text=parts[1], font=error_inline_font))
                            # 分隔符
                            rich_text.append(TextBlock(text="/", font=default_inline_font))
                            # 未回答率（灰色）
                            rich_text.append(TextBlock(text=parts[2], font=miss_inline_font))
                            
                            cell.value = rich_text
                            cell.font = default_font  # 设置基础字体为微软雅黑
                    except (ValueError, IndexError) as e:
                        print(f"Warning: Could not parse value '{cell.value}' for formatting")
                        continue
        
        # 设置不同的背景颜色
        overview_fill = PatternFill(start_color=COLORS['overview_total'], 
                                  end_color=COLORS['overview_total'], 
                                  fill_type='solid')
        category_fill = PatternFill(start_color=COLORS['category_total'], 
                                  end_color=COLORS['category_total'], 
                                  fill_type='solid')
        final_total_fill = PatternFill(start_color=COLORS['final_total'], 
                                     end_color=COLORS['final_total'], 
                                     fill_type='solid')
        
        # 为不同类型的汇总行添加背景色
        for row in worksheet.iter_rows(min_row=2):
            cell_value = str(row[0].value) if row[0].value else ''
            if 'Overall Total' in cell_value:  # 最终汇总行
                for cell in row:
                    cell.fill = final_total_fill
            elif 'Total' in cell_value:  # Overview汇总行
                for cell in row:
                    cell.fill = overview_fill
            else:
                category_cell = str(row[1].value) if row[1].value else ''
                if 'Total' in category_cell:  # Category汇总行
                    for cell in row:
                        cell.fill = category_fill
        
        # 合并单元格
        current_overview = None
        current_category = None
        overview_start_row = 2  # Excel的行号从1开始，标题占据第1行
        category_start_row = 2
        
        for idx, row in df.iterrows():
            current_row = idx + 2  # 加2是因为Excel行号从1开始且有标题行
            
            # 处理Overview Category的合并
            if row['Overview Category'] != current_overview:
                if current_overview is not None and 'Total' not in current_overview:
                    # 只有当不是汇总行且有多行需要合并时才合并
                    if current_row - overview_start_row > 1:
                        worksheet.merge_cells(f'A{overview_start_row}:A{current_row-1}')
                current_overview = row['Overview Category']
                overview_start_row = current_row
            
            # 处理Category的合并
            if row['Category'] != current_category:
                if current_category is not None and current_category != '':
                    # 只有当不是空的category且有多行需要合并时才合并
                    if current_row - category_start_row > 1:
                        worksheet.merge_cells(f'B{category_start_row}:B{current_row-1}')
                current_category = row['Category']
                category_start_row = current_row
        
        # 处理最后一组数据的合并
        last_row = len(df) + 1
        if current_overview is not None and 'Total' not in current_overview:
            if last_row - overview_start_row > 1:
                worksheet.merge_cells(f'A{overview_start_row}:A{last_row}')
        if current_category is not None and current_category != '':
            if last_row - category_start_row > 1:
                worksheet.merge_cells(f'B{category_start_row}:B{last_row}')
        
        # 调整列宽以适应新的格式
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # 添加说明行
        last_row = len(df) + 3  # 空一行再添加说明
        
        # 写入说明内容
        parts = [
            ("数值说明: ", None, None),
            ("正确率", COLORS['acc'], None),
            ("/", None, None),
            ("错误率", COLORS['error'], None),
            ("/", None, None),
            ("未回答率", COLORS['miss'], None),
            (" 格式中，", None, None),
            ("黄底", COLORS['overview_total'], None),
            ("表示大类汇总，", None, None),
            ("浅黄底", COLORS['category_total'], None),
            ("表示小类汇总，", None, None),
            ("红底", COLORS['final_total'], None),
            ("表示总计", None, None)
        ]
        
        # 合并单元格
        worksheet.merge_cells(f'A{last_row}:M{last_row}')
        merged_cell = worksheet.cell(row=last_row, column=1)
        
        # 构建富文本说明
        rich_text = CellRichText()
        for text, color, fill_color in parts:
            if color:
                rich_text.append(TextBlock(text=text, font=InlineFont(color=color)))
            else:
                rich_text.append(TextBlock(text=text, font=default_inline_font))
        
        # 设置单元格值和格式
        merged_cell.value = rich_text
        merged_cell.font = default_font
        merged_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 设置背景色
        if fill_color:
            merged_cell.fill = PatternFill(start_color=fill_color, 
                                         end_color=fill_color, 
                                         fill_type='solid')
        
        # 设置行高
        worksheet.row_dimensions[last_row].height = 30

    print(f"修正完成，Excel文件生成成功：{save_path}")

def format_cell_value(stats):
    """格式化单元格内容，返回带有acc/error/miss的字符串"""
    total = stats['total']
    if total == 0:
        return '0%/0%/0%'
    
    acc = stats['correct'] / total
    error = stats['error'] / total
    miss = stats['miss'] / total
    
    return f"{acc:.1%}/{error:.1%}/{miss:.1%}"

def main(args):
    # 创建PrettyTable用于显示结果
    results = PrettyTable()
    results.field_names = ["Model", "Split", "Mode", "Accuracy", "Errors", "Miss"]
    
    # 收集所有数据
    model_results = {}
    hierarchy_stats = {
        'overview': {},
        'category': {},
        'subcategory': {}
    }
    
    # 确定要处理的文件
    if args.evaluate_all:
        files = sorted([f for f in os.listdir(args.output_dir) if f.endswith('.jsonl')])
        output_suffix = args.split + '_all_models'
    else:
        # 处理指定的模型和模式
        if not isinstance(args.model_name, list):
            args.model_name = [args.model_name]
        
        files = []
        for model in args.model_name:
            for mode in args.mode:
                file_name = f"{model}_{args.split}_{mode}.jsonl"
                if os.path.exists(os.path.join(args.output_dir, file_name)):
                    files.append(file_name)
        output_suffix = args.split + '_' + '_'.join(args.model_name)
    
    # 简化的结果结构
    final_results = {}
    
    # 处理文件
    for file_name in tqdm(files, desc="处理文件"):
        if args.split == '':
            model_name, split, mode = file_name.split('_')
        else:
            if args.split in file_name:
                model_name, mode = file_name.split(f'_{args.split}_')
                split = args.split
            else:
                continue
        mode = mode.replace('.jsonl', '')
        
        file_path = os.path.join(args.output_dir, file_name)
        accuracy, errors, miss, stats = calculate_accuracy(file_path, args.save_dir)
        
        # 添加到PrettyTable
        results.add_row([model_name, split, mode, f"{accuracy:.2%}", f"{errors:.2%}", f"{miss:.2%}"])
        
        # 保存统计数据用于生成Excel报告
        if model_name not in model_results:
            model_results[model_name] = {}
        model_results[model_name][mode] = stats
        
        # 更新层级统计
        for level in ['overview', 'category', 'subcategory']:
            for key, data in stats[level].items():
                if key not in hierarchy_stats[level]:
                    hierarchy_stats[level][key] = data.copy()
                    hierarchy_stats[level][key]['model_stats'] = {}
                
                if model_name not in hierarchy_stats[level][key]['model_stats']:
                    hierarchy_stats[level][key]['model_stats'][model_name] = {}
                hierarchy_stats[level][key]['model_stats'][model_name][mode] = data
        
        # 初始化模型结果
        if model_name not in final_results:
            final_results[model_name] = {
                "zero-shot": {
                    "accuracy": 0,
                    "errors": 0,
                    "miss": 0,
                    "categories": {}
                }
            }
        
        # 更新总体指标
        final_results[model_name][mode]["accuracy"] = accuracy
        final_results[model_name][mode]["errors"] = errors 
        final_results[model_name][mode]["miss"] = miss
        
        # 更新所有层级的分类统计
        categories_dict = final_results[model_name][mode]["categories"]
        
        # 遍历stats中的所有层级
        for hierarchy_level, level_stats in stats.items():
            if hierarchy_level == "overview":
                continue  # 跳过总览数据
                
            # 对于每个层级的每个类别
            for category, category_stats in level_stats.items():
                if category not in categories_dict:
                    categories_dict[category] = {
                        "correct": category_stats["correct"],
                        "total": category_stats["total"],
                        "error": category_stats["error"],
                        "miss": category_stats["miss"]
                    }
    
    # 打印总体结果
    print(results)
    
    # 生成Excel报告
    if args.excel_output:
        # 根据模式生成不同的文件名
        output_file = os.path.join(
            args.save_dir, 
            f'results_{output_suffix}.xlsx'
        )
        create_excel_report_from_stats(
            model_results,
            hierarchy_stats,
            output_file
        )
    
    # 保存JSON结果
    if args.json_output:
        json_output_file = os.path.join(
            args.save_dir, 
            f'results_{output_suffix}.json'
        )
        with open(json_output_file, 'w', encoding='utf-8') as f:
            json.dump(final_results, f, ensure_ascii=False, indent=2)
        print(f"JSON结果已保存至: {json_output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Calculate accuracy for different modes and splits.")
    parser.add_argument('--model_name', type=str, nargs='+', default=[], help='Model names to use')
    # parser.add_argument('--split', type=str, default='GPQA-data-processed-by-lnn-20241211-version2', help='Data split to use')
    parser.add_argument('--split', type=str, default='toprp-with-confusion-options', help='Data split to use')
    parser.add_argument('--mode', nargs='+', default=['zero-shot'], help='Modes to use for data loading')
    parser.add_argument('--output_dir', type=str, default='results/gpqa', help='Directory to read result files from')
    parser.add_argument('--save_dir', type=str, default='results_with_status/gpqa', help='Directory to save result files with status')
    parser.add_argument('--evaluate_all', action='store_true', help='Evaluate all files in the output directory')
    parser.add_argument('--excel_output', action='store_true', help='Generate Excel report with category-wise results')
    parser.add_argument('--json_output', action='store_true', help='Generate JSON file with detailed results')
    
    args = parser.parse_args()
    
    # 验证参数
    if not args.evaluate_all and not args.model_name:
        parser.error("Either --evaluate_all or --model_name must be specified")
    
    main(args)
